#!/usr/bin/env python3
"""解析销售小票Excel (45)，统计每人6月业绩"""
import openpyxl
import re
from collections import defaultdict

FILE = '/Users/a86137/Downloads/销售小票查询 (45).XLSX'

# 缩写映射（大小写不敏感）
NAME_MAP = {
    'lrt': '李若彤', 'kxy': '孔祥宇', 'cxy': '陈昕媛', 'wjy': '王靳毓', 'tjl': '田佳乐',
    'cc': '迟骋', 'dqy': '邓奇缘', 'yzh': '杨子豪', 'wyl': '王雅澜',
    'hqy': '何秋烨', 'zky': '朱凯赟', 'wly': '王龙宇', 'gyh': '龚赟昊',
}

def match_name(remark):
    """从备注中匹配兼职缩写"""
    if not remark:
        return None
    remark_lower = str(remark).lower().strip()
    # 去掉空格再匹配
    remark_nospace = remark_lower.replace(' ', '')
    for abbr, name in NAME_MAP.items():
        if abbr in remark_nospace:
            return name
    return None

def classify(category):
    """品类分类逻辑（已修正）"""
    if not category:
        return '其他'
    cat = str(category).strip()
    if any(x in cat for x in ['鞋', '靴']):
        return '鞋履'
    if any(x in cat for x in ['服', '衣', '裤', '裙', 'T恤', '卫衣', '外套', '茄克', 'POLO', '衫', '背心']):
        return '服装'
    if any(x in cat for x in ['配', '袜', '帽', '包', '围巾', '手套', '袋']):
        return '配件'
    return '其他'

# 读取Excel
wb = openpyxl.load_workbook(FILE, data_only=True)
ws = wb['Sheet1']

# 行2是表头（行1是大标题"销售小票查询"）
headers = {}
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=2, column=col).value
    if val:
        headers[str(val).strip()] = col

print("=== 表头列名 ===")
for h, c in headers.items():
    print(f"  列{c}: {h}")

# 精确列索引
COL_TYPE = 1       # 小票类型
COL_ORDER = 3      # 小票号
COL_SALESTYPE = 4  # 销售类型
COL_DATE = 5       # 销售日期
COL_QTY = 10       # 销售数量
COL_AMOUNT = 13    # 实际零售金额
COL_REMARK = 20    # 备注
COL_BIGCAT = 24    # 大类（鞋履/服装/配件）
COL_PRODCAT = 25   # 产品类别

print(f"\n备注列={COL_REMARK}, 金额列={COL_AMOUNT}, 数量列={COL_QTY}")
print(f"大类列={COL_BIGCAT}, 产品类别列={COL_PRODCAT}")

# 统计数据
staff_data = defaultdict(lambda: {'sales': 0, 'qty': 0, 'cats': defaultdict(lambda: {'sales': 0, 'qty': 0})})
refunds = []  # 退换货记录
unmatched = []

# 数据从行3开始
for row in range(3, ws.max_row + 1):
    remark = ws.cell(row=row, column=COL_REMARK).value
    amount = ws.cell(row=row, column=COL_AMOUNT).value
    qty = ws.cell(row=row, column=COL_QTY).value
    bigcat = ws.cell(row=row, column=COL_BIGCAT).value  # 大类：鞋履/服装/配件
    prodcat = ws.cell(row=row, column=COL_PRODCAT).value  # 产品类别
    rtype = ws.cell(row=row, column=COL_TYPE).value  # 小票类型
    order_no = ws.cell(row=row, column=COL_ORDER).value
    
    if not remark:
        continue
    
    # 金额处理
    try:
        amount = float(amount) if amount else 0
    except:
        amount = 0
    
    try:
        qty = int(float(qty)) if qty else 1
    except:
        qty = 1
    
    name = match_name(remark)
    
    # 分类：优先用"大类"列，如果没有则用产品类别推断
    if bigcat and str(bigcat).strip() in ['鞋履', '服装', '配件']:
        cat = str(bigcat).strip()
    else:
        cat = classify(prodcat)
    
    # 检查是否退换货
    rtype_str = str(rtype) if rtype else ''
    remark_str = str(remark)
    
    if '退货' in rtype_str or '退货' in remark_str or '全退' in rtype_str:
        refunds.append({
            'order': str(order_no) if order_no else '',
            'amount': amount,
            'qty': qty,
            'remark': remark_str,
            'category': str(bigcat) if bigcat else '',
        })
        continue
    
    if '换货' in rtype_str or '换货' in remark_str:
        # 换货：净差额处理（金额可为正可为负）
        name = match_name(remark)
        if name:
            staff_data[name]['sales'] += amount
            staff_data[name]['qty'] += qty
            staff_data[name]['cats'][cat]['sales'] += amount
            staff_data[name]['cats'][cat]['qty'] += qty
        continue
    
    if name:
        staff_data[name]['sales'] += amount
        staff_data[name]['qty'] += qty
        staff_data[name]['cats'][cat]['sales'] += amount
        staff_data[name]['cats'][cat]['qty'] += qty
    else:
        unmatched.append(remark_str)

print("\n=== 各人业绩统计 ===")
total_sales = 0
total_qty = 0
for name in ['陈昕媛','田佳乐','迟骋','王靳毓','朱凯赟','孔祥宇','邓奇缘','杨子豪','王雅澜','李若彤','王龙宇','何秋烨','龚赟昊']:
    d = staff_data.get(name, {'sales': 0, 'qty': 0, 'cats': {}})
    sales = round(d['sales'])
    qty = d['qty']
    total_sales += sales
    total_qty += qty
    
    # 品类占比
    cat_str = ''
    if d['cats']:
        cats_sorted = sorted(d['cats'].items(), key=lambda x: -x[1]['sales'])
        parts = []
        for cn, cd in cats_sorted:
            if sales > 0:
                pct = round(cd['sales'] / sales * 100, 1)
                parts.append(f"{cn} {pct}%")
        cat_str = ' / '.join(parts)
    
    print(f"  {name}: ¥{sales:,} ({qty}件) [{cat_str}]")

print(f"\n总计: ¥{total_sales:,} ({total_qty}件)")

print(f"\n=== 退换货记录 ({len(refunds)}条) ===")
for r in refunds[:10]:
    print(f"  {r['order']}: ¥{r['amount']} ({r['qty']}件) - {r['remark'][:30]}")

if len(refunds) > 10:
    print(f"  ... 共{len(refunds)}条")

print(f"\n=== 未匹配备注 ({len(unmatched)}条) ===")
for u in unmatched[:10]:
    print(f"  {u}")

# 输出categories字符串供更新
print("\n=== categories字符串 ===")
for name in ['陈昕媛','田佳乐','迟骋','王靳毓','朱凯赟','孔祥宇','邓奇缘','杨子豪','王雅澜','李若彤','王龙宇','何秋烨','龚赟昊']:
    d = staff_data.get(name, {'sales': 0, 'qty': 0, 'cats': {}})
    sales = round(d['sales'])
    if d['cats'] and sales > 0:
        cats_sorted = sorted(d['cats'].items(), key=lambda x: -x[1]['sales'])
        parts = []
        for cn, cd in cats_sorted:
            pct = round(cd['sales'] / sales * 100, 1)
            parts.append(f"{cn} {pct}%")
        print(f"  {name}: categories: '{' / '.join(parts)}'")
