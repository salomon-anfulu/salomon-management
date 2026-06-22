#!/usr/bin/env python3
"""Parse 销售小票查询(42).XLSX and compute per-person performance data.
Key: use 结算金额(col 13) which has correct sign for returns/exchanges.
Trace returns back to original sales staff via ticket number in remark.
"""
import openpyxl
import json
import re
from collections import defaultdict

XLSX_PATH = "/Users/a86137/Downloads/销售小票查询 (42).XLSX"

# 缩写 -> 全名映射（大小写不敏感）
NAME_MAP = {
    'lrt': '李若彤', 'kxy': '孔祥宇', 'cxy': '陈昕媛', 'wjy': '王靳毓', 'tjl': '田佳乐',
    'cc': '迟骋', 'dqy': '邓奇缘', 'yzh': '杨子豪', 'wyl': '王雅澜',
    'hqy': '何秋烨', 'zky': '朱凯赟', 'wly': '王龙宇', 'gyh': '龚赟昊',
}

# Latest work hours from linggong attendance (2026-06-22)
WORK_HOURS = {
    '陈昕媛': 102.5, '杨子豪': 78.0, '李若彤': 90.0, '孔祥宇': 86.0,
    '王雅澜': 78.0, '龚赟昊': 78.0, '邓奇缘': 83.0, '王龙宇': 58.5,
    '何秋烨': 74.0, '田佳乐': 78.5, '朱凯赟': 75.0, '王靳毓': 74.5,
    '迟骋': 78.0,
}

WORK_DAYS = {
    '陈昕媛': 14, '杨子豪': 10, '李若彤': 13, '孔祥宇': 13,
    '王雅澜': 11, '龚赟昊': 11, '邓奇缘': 12, '王龙宇': 8,
    '何秋烨': 12, '田佳乐': 11, '朱凯赟': 10, '王靳毓': 11,
    '迟骋': 12,
}

def match_name(remark):
    if not remark:
        return None
    remark_lower = str(remark).lower().strip()
    for abbr, name in NAME_MAP.items():
        if abbr in remark_lower:
            return name
    return None

def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb.active
    
    # Header is row 2
    header_row_idx = 2
    headers = [cell.value for cell in ws[header_row_idx]]
    
    # Column indices (0-based)
    COL_TYPE = 0       # 小票类型: 销售/全退货/换货/合计
    COL_TICKET = 2     # 小票号
    COL_DATE = 4       # 销售日期
    COL_QTY = 9        # 销售数量
    COL_AMOUNT = 13    # 结算金额 (has correct sign for returns)
    COL_REMARK = 19    # 备注
    COL_CATEGORY = 24  # 产品类别(大类)
    
    # Parse all rows, skip 合计
    all_rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        row_type = str(row[COL_TYPE] or '').strip()
        if not row_type or row_type == '合计':
            continue
        if not row[COL_TICKET]:
            continue
        
        amount = row[COL_AMOUNT] or 0
        qty = row[COL_QTY] or 0
        remark = str(row[COL_REMARK] or '').strip()
        ticket = str(row[COL_TICKET] or '').strip()
        category = str(row[COL_CATEGORY] or '其他').strip()
        
        all_rows.append({
            'type': row_type,
            'ticket': ticket,
            'amount': amount,
            'qty': qty,
            'remark': remark,
            'category': category,
        })
    
    print(f"Total parsed rows: {len(all_rows)}")
    
    # Build ticket -> name mapping from 销售 rows
    ticket_to_name = {}
    for r in all_rows:
        if r['type'] == '销售':
            name = match_name(r['remark'])
            if name:
                ticket_to_name[r['ticket']] = name
    
    # Aggregate per person
    data = defaultdict(lambda: {
        'sales': 0, 'qty': 0, 'tickets': set(),
        'categories': defaultdict(lambda: {'sales': 0, 'qty': 0})
    })
    
    untraced_returns = []
    
    for r in all_rows:
        name = None
        amount = r['amount']
        qty = r['qty']
        ticket = r['ticket']
        remark = r['remark']
        category = r['category']
        row_type = r['type']
        
        if row_type == '销售':
            name = match_name(remark)
        elif row_type in ('全退货', '换货'):
            # Try to trace original ticket from remark
            m = re.search(r'小票号:([^,)]+)', remark)
            if m:
                orig_ticket = m.group(1).strip()
                name = ticket_to_name.get(orig_ticket)
                if not name:
                    for t, n in ticket_to_name.items():
                        if orig_ticket in t or t in orig_ticket:
                            name = n
                            break
            if not name:
                name = match_name(remark)
            
            if not name:
                untraced_returns.append(r)
                continue
        
        if not name:
            continue
        
        # Normalize category
        if any(x in category for x in ['鞋', '靴']):
            cat = '鞋履'
        elif any(x in category for x in ['服', '衣', '裤', '裙', 'T恤', '卫衣', '外套']):
            cat = '服装'
        elif any(x in category for x in ['配', '袜', '帽', '包', '围巾', '手套']):
            cat = '配件'
        else:
            cat = '其他'
        
        data[name]['sales'] += amount
        data[name]['qty'] += qty
        if row_type == '销售':
            data[name]['tickets'].add(ticket)
        data[name]['categories'][cat]['sales'] += amount
        data[name]['categories'][cat]['qty'] += qty
    
    if untraced_returns:
        print(f"\n Untraced returns/exchanges ({len(untraced_returns)}):")
        for r in untraced_returns:
            print(f"  type={r['type']}, ticket={r['ticket']}, amount={r['amount']}, remark={r['remark']}")
    
    # Calculate total
    total_sales = sum(d['sales'] for d in data.values())
    total_hours = sum(WORK_HOURS.values())
    avg_hourly = total_sales / total_hours if total_hours > 0 else 0
    
    # Build records sorted by sales desc
    records = []
    for name, d in sorted(data.items(), key=lambda x: -x[1]['sales']):
        sales = round(d['sales'])
        qty = d['qty']
        tickets = len(d['tickets'])
        avg_price = round(sales / qty) if qty > 0 else 0
        work_hours = WORK_HOURS.get(name, 50.0)
        work_days = WORK_DAYS.get(name, 7)
        hourly_output = round(sales / work_hours) if work_hours > 0 else 0
        sales_share = round(sales / total_sales, 3) if total_sales > 0 else 0
        efficiency = round((hourly_output - avg_hourly) / avg_hourly, 4) if avg_hourly > 0 else 0
        
        cats = {}
        for cat_name, cd in sorted(d['categories'].items(), key=lambda x: -x[1]['sales']):
            cat_sales = round(cd['sales'])
            cat_qty = cd['qty']
            pct = round(cat_sales / sales * 100, 1) if sales > 0 else 0
            if cat_sales != 0:
                cats[cat_name] = {'sales': cat_sales, 'qty': cat_qty, 'pct': pct}
        
        records.append({
            'name': name,
            'sales': sales,
            'qty': qty,
            'tickets': tickets,
            'avgPrice': avg_price,
            'workHours': work_hours,
            'workDays': work_days,
            'hourlyOutput': hourly_output,
            'efficiency': efficiency,
            'salesShare': sales_share,
            'categories': cats
        })
    
    # Calculate avg UPT
    total_qty = sum(r['qty'] for r in records)
    total_tickets = sum(r['tickets'] for r in records)
    avg_upt = round(total_qty / total_tickets, 2) if total_tickets > 0 else 0
    
    result = {
        'totalSales': total_sales,
        'avgUPT': avg_upt,
        'avgHourlyOutput': round(avg_hourly, 1),
        'records': records,
    }
    
    print(f"\n=== RESULTS ===")
    print(f"Total Sales: {total_sales:,}")
    print(f"Avg UPT: {avg_upt}")
    print(f"Avg Hourly Output: {round(avg_hourly, 1)}")
    print(f"People: {len(records)}")
    print(f"{'Name':<8} {'Sales':>8} {'Qty':>4} {'Tkt':>4} {'Hourly':>7} {'Share':>6} {'Hours':>6} {'Days':>5}")
    for r in records:
        print(f"{r['name']:<8} {r['sales']:>7,} {r['qty']:>4} {r['tickets']:>4} {r['hourlyOutput']:>5} {r['salesShare']*100:>5.1f}% {r['workHours']:>5.1f}h {r['workDays']:>4}d")
    
    with open('/Users/a86137/Desktop/兼职/安福路兼职管理系统/data/june_perf_42.json', 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f"\nSaved to data/june_perf_42.json")

if __name__ == '__main__':
    main()
